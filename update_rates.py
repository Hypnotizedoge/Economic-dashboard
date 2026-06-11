import urllib.request
import json
import csv
import datetime
import os
import sys
import pandas as pd
from pathlib import Path

# Ensure script directory is in sys.path
sys.path.append(str(Path(__file__).parent.resolve()))

API_URL = "https://api.bnm.gov.my/public/exchange-rate"
HEADERS = {
    "Accept": "application/vnd.BNM.API.v1+json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}
YAHOO_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

CSV_FILE = "FXrate.csv"

# The standard currency order in the CSV header
CURRENCIES = [
    "USD", "EUR", "JPY", "GBP", "SGD", "AUD", "CAD", "CNY", "CHF", "THB",
    "IDR", "VND", "KHR", "MMK", "PHP", "BND", "KRW", "HKD", "TWD", "INR",
    "PKR", "NPR", "SAR", "AED", "EGP", "NZD", "XDR"
]

def format_rate(rate_val):
    if rate_val is None:
        return ""
    # Format to 8 decimal places and strip trailing zeros/dot to avoid floating point anomalies
    s = f"{rate_val:.8f}".rstrip('0').rstrip('.')
    return s

def _read_existing_dates() -> set:
    """Read all dates already present in FXrate.csv."""
    existing = set()
    if os.path.exists(CSV_FILE):
        with open(CSV_FILE, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)  # skip header
            for row in reader:
                if row:
                    existing.add(row[0].strip())
    return existing


def _find_last_csv_date(existing_dates: set) -> datetime.date | None:
    """Parse all D/M/YYYY date strings and return the most recent as a date."""
    parsed = []
    for d in existing_dates:
        try:
            parsed.append(datetime.datetime.strptime(d, "%d/%m/%Y").date())
        except ValueError:
            # Also try single-digit day/month format (e.g. 1/6/2025)
            try:
                parts = d.split("/")
                parsed.append(datetime.date(int(parts[2]), int(parts[1]), int(parts[0])))
            except (ValueError, IndexError):
                continue
    return max(parsed) if parsed else None


def _fetch_rates_for_date(target_date: datetime.date) -> list | None:
    """Fetch exchange-rate records from BNM API for a specific date.
    Returns the list of currency records, or None on failure / no data.
    """
    date_str = target_date.strftime("%Y-%m-%d")
    url = f"{API_URL}?date={date_str}"
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=15) as response:
            if response.getcode() != 200:
                return None
            body = response.read().decode("utf-8")
            data = json.loads(body)
            records = data.get("data", [])
            return records if records else None
    except Exception:
        return None


def _build_csv_row(csv_date_str: str, records: list) -> list:
    """Build a CSV row from BNM API records for a given date."""
    rate_map = {}
    for record in records:
        code = record.get("currency_code")
        if code:
            rate_map[code] = record

    new_row = [csv_date_str, "middle"]
    for col_currency in CURRENCIES:
        lookup_code = "SDR" if col_currency == "XDR" else col_currency
        entry = rate_map.get(lookup_code)

        rate_str = ""
        if entry:
            rate_info = entry.get("rate", {})
            middle_rate = rate_info.get("middle_rate")
            unit = entry.get("unit", 1)
            if middle_rate is not None:
                try:
                    normalized_rate = float(middle_rate) / float(unit)
                    rate_str = format_rate(normalized_rate)
                except (ValueError, TypeError) as e:
                    print(f"  Error processing rate for {lookup_code}: {e}")

        new_row.append(rate_str)
    return new_row


def _ensure_trailing_newline():
    """Make sure FXrate.csv ends with a newline before appending."""
    if os.path.exists(CSV_FILE):
        with open(CSV_FILE, "rb+") as f:
            f.seek(0, 2)
            if f.tell() > 0:
                f.seek(-1, 2)
                if f.read(1) != b"\n":
                    f.write(b"\n")


def update_exchange_rates():
    import time

    print(f"Exchange rate source: BNM Open API ({API_URL})")

    # 1. Read existing dates and find the last one
    existing_dates = _read_existing_dates()
    last_date = _find_last_csv_date(existing_dates)
    today = datetime.date.today()

    if last_date is None:
        # No data at all — just fetch today's rate
        print("No existing dates found in CSV. Fetching today's rate only.")
        dates_to_fetch = [today]
    elif last_date >= today:
        print(f"CSV is already up-to-date (last entry: {last_date}). Nothing to fetch.")
        return
    else:
        # Build list of all missing dates from (last_date + 1) to today
        gap_days = (today - last_date).days
        dates_to_fetch = [last_date + datetime.timedelta(days=i) for i in range(1, gap_days + 1)]
        print(f"Last CSV date: {last_date}  |  Today: {today}  |  {len(dates_to_fetch)} day(s) to backfill")

    # 2. Fetch and append each missing date
    appended = 0
    skipped = 0

    _ensure_trailing_newline()

    for target_date in dates_to_fetch:
        csv_date_str = f"{target_date.day}/{target_date.month}/{target_date.year}"

        # Skip if already in CSV (safety check)
        if csv_date_str in existing_dates:
            continue

        records = _fetch_rates_for_date(target_date)
        if records is None:
            # No data for this date (weekend / public holiday / API issue)
            skipped += 1
            print(f"  {target_date}  - no data (weekend/holiday), skipped")
            continue

        new_row = _build_csv_row(csv_date_str, records)

        with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(new_row)

        existing_dates.add(csv_date_str)
        appended += 1
        print(f"  {target_date}  - appended [OK]")

        # Polite delay between API requests (avoid hammering the server)
        if len(dates_to_fetch) > 1:
            time.sleep(0.5)

    print(f"Exchange rates done: {appended} day(s) appended, {skipped} day(s) skipped (no data).")

def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def update_money_supply():
    excel_file = "Money supply.xlsx"
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found.")
        return

    print(f"Loading {excel_file}...")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["1.3"]
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    # Find the last date in Excel
    last_row = ws.max_row
    current_year = None
    last_year = None
    last_month = None

    for row in range(10, last_row + 1):
        val_a = ws.cell(row=row, column=1).value
        val_b = ws.cell(row=row, column=2).value
        if val_a is not None:
            try:
                current_year = int(val_a)
            except:
                pass
        if val_b is not None:
            try:
                m = int(val_b)
                if current_year is not None:
                    last_year = current_year
                    last_month = m
            except:
                pass

    if last_year is None or last_month is None:
        print("Could not find the last date in the Excel file.")
        return

    print(f"Last date in Excel: Month {last_month}, Year {last_year}")

    # Fetch data from BNM API for last_year up to current system year
    current_system_year = datetime.datetime.now().year
    new_records = []

    for year in range(last_year, current_system_year + 1):
        url = f"https://api.bnm.gov.my/public/msb/1.3/year/{year}"
        print(f"Fetching money supply data from BNM API for year {year}: {url}")
        req = urllib.request.Request(url, headers=HEADERS)
        try:
            with urllib.request.urlopen(req, timeout=15) as response:
                status = response.getcode()
                if status != 200:
                    print(f"Failed to fetch year {year}. HTTP Status: {status}")
                    continue
                body = response.read().decode('utf-8')
                data = json.loads(body)
                records = data.get('data', [])
                new_records.extend(records)
        except Exception as e:
            print(f"Error fetching year {year}: {e}")
            continue

    # Filter and sort new records
    records_to_append = []
    for r in new_records:
        try:
            y = int(r.get('year_dt', 0))
            m = int(r.get('month_dt', 0))
        except (ValueError, TypeError):
            continue
            
        if y > last_year or (y == last_year and m > last_month):
            records_to_append.append((y, m, r))

    # Sort by year, then month
    records_to_append.sort(key=lambda x: (x[0], x[1]))

    if not records_to_append:
        print("No new money supply data found in the API compared to Excel. Skipping append.")
        return

    # Append to Excel
    for y, m, r in records_to_append:
        next_row = ws.max_row + 1
        # Col 1 (A): Year (only for January)
        ws.cell(row=next_row, column=1, value=y if m == 1 else None)
        # Col 2 (B): Month
        ws.cell(row=next_row, column=2, value=m)
        # Col 3 (C): None
        ws.cell(row=next_row, column=3, value=None)
        
        # Col 4-16 (D-P)
        ws.cell(row=next_row, column=4, value=safe_float(r.get('m_thr_tot')))
        ws.cell(row=next_row, column=5, value=safe_float(r.get('m_two_tot')))
        ws.cell(row=next_row, column=6, value=safe_float(r.get('m_one_tot')))
        ws.cell(row=next_row, column=7, value=safe_float(r.get('cur_in_cir')))
        ws.cell(row=next_row, column=8, value=safe_float(r.get('dem_dep')))
        ws.cell(row=next_row, column=9, value=safe_float(r.get('tot')))
        ws.cell(row=next_row, column=10, value=safe_float(r.get('nqm_sav_dep')))
        ws.cell(row=next_row, column=11, value=safe_float(r.get('nqm_fix_dep')))
        ws.cell(row=next_row, column=12, value=safe_float(r.get('nqm_nid')))
        ws.cell(row=next_row, column=13, value=safe_float(r.get('nqm_rep')))
        ws.cell(row=next_row, column=14, value=safe_float(r.get('nqm_for_cur_dep')))
        ws.cell(row=next_row, column=15, value=safe_float(r.get('nqm_oth_dep')))
        ws.cell(row=next_row, column=16, value=safe_float(r.get('dep_pla_oth_ban_ins')))
        
        print(f"Appended month {m}, year {y} to {excel_file}.")

    # Save workbook
    try:
        wb.save(excel_file)
        print(f"Successfully saved updates to {excel_file}.")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

def clean_price(val):
    if pd.isna(val) or str(val).strip() in ("", "-"):
        return float("nan")
    return float(str(val).replace(",", "").replace('"', '').strip())

def format_price(val):
    if pd.isna(val) or val is None:
        return ""
    try:
        return f"{float(val):,.2f}"
    except:
        return str(val)

def format_volume(vol_val):
    if vol_val is None or vol_val == 0 or pd.isna(vol_val):
        return "-"
    try:
        val = float(vol_val)
        if val >= 1e9:
            return f"{val/1e9:.2f}B"
        if val >= 1e6:
            return f"{val/1e6:.2f}M"
        if val >= 1e3:
            return f"{val/1e3:.2f}K"
        return str(int(val))
    except:
        return "-"

def update_stock_index_csv(filename: str, ticker: str):
    print(f"Updating stock index: {filename} ({ticker})")
    csv_path = Path(__file__).parent / "Stock indices" / filename
    if not csv_path.exists():
        print(f"  Error: {filename} not found.")
        return
        
    # Load existing CSV
    df_existing = pd.read_csv(csv_path)
    df_existing.columns = df_existing.columns.str.strip().str.replace('"', '')
    
    # Parse existing Dates
    df_existing["ParsedDate"] = pd.to_datetime(df_existing["Date"], format="%m/%d/%Y", errors="coerce")
    # Drop rows with invalid dates
    df_existing = df_existing.dropna(subset=["ParsedDate"])
    
    # Find last date
    last_date = df_existing["ParsedDate"].max()
    today = datetime.date.today()
    
    # We want to fetch from 5 days before last_date to ensure overlap and correct change percentage calculation
    start_date = last_date.date() - datetime.timedelta(days=5)
    
    period1 = int(datetime.datetime.combine(start_date, datetime.time.min).timestamp())
    period2 = int(datetime.datetime.combine(today, datetime.time.max).timestamp())
    
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&period1={period1}&period2={period2}"
    req = urllib.request.Request(url, headers=YAHOO_HEADERS)
    
    try:
        with urllib.request.urlopen(req, timeout=15) as response:
            if response.getcode() != 200:
                print(f"  Failed to fetch data for {ticker} (status {response.getcode()})")
                return
            body = response.read().decode("utf-8")
            data = json.loads(body)
            result = data["chart"]["result"][0]
            timestamps = result.get("timestamp", [])
            quote = result["indicators"]["quote"][0]
    except Exception as e:
        print(f"  Error fetching {ticker}: {e}")
        return
        
    if not timestamps:
        print(f"  No new data found for {ticker}.")
        return
        
    new_rows = []
    for i, ts in enumerate(timestamps):
        date_val = datetime.datetime.fromtimestamp(ts).date()
        o = quote["open"][i]
        h = quote["high"][i]
        l = quote["low"][i]
        c = quote["close"][i]
        v = quote["volume"][i] if ("volume" in quote and quote["volume"]) else 0
        
        if o is None or h is None or l is None or c is None:
            continue
            
        new_rows.append({
            "ParsedDate": pd.to_datetime(date_val),
            "Price": float(c),
            "Open": float(o),
            "High": float(h),
            "Low": float(l),
            "Vol.": v
        })
        
    if not new_rows:
        print(f"  No valid trading days returned for {ticker}.")
        return
        
    df_new = pd.DataFrame(new_rows)
    
    # Merge existing and new
    # For existing, convert Price/Open/High/Low to float to clean and align
    for col in ["Price", "Open", "High", "Low"]:
        df_existing[col] = df_existing[col].apply(clean_price)
    
    # Drop duplicates by ParsedDate, keeping the new data
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    df_combined = df_combined.drop_duplicates(subset=["ParsedDate"], keep="last")
    
    # Sort chronologically to compute Change % properly
    df_combined = df_combined.sort_values("ParsedDate").reset_index(drop=True)
    
    # Recalculate Change %
    df_combined["Change_Raw"] = df_combined["Price"].pct_change() * 100
    
    # Format columns for CSV
    df_combined["Date"] = df_combined["ParsedDate"].dt.strftime("%m/%d/%Y")
    df_combined["Price"] = df_combined["Price"].apply(format_price)
    df_combined["Open"] = df_combined["Open"].apply(format_price)
    df_combined["High"] = df_combined["High"].apply(format_price)
    df_combined["Low"] = df_combined["Low"].apply(format_price)
    
    # Format Vol.
    def format_vol_col(row):
        val = row["Vol."]
        if pd.isna(val) or val is None or val == "":
            return "-"
        try:
            if isinstance(val, str) and any(c in val for c in ["M", "B", "K", "-"]):
                return val
            val_f = float(val)
            if val_f == 0:
                return "-"
            if val_f >= 1e9:
                return f"{val_f/1e9:.2f}B"
            if val_f >= 1e6:
                return f"{val_f/1e6:.2f}M"
            if val_f >= 1e3:
                return f"{val_f/1e3:.2f}K"
            return str(int(val_f))
        except:
            return str(val)
            
    df_combined["Vol."] = df_combined.apply(format_vol_col, axis=1)
    
    def format_change_col(row):
        val = row["Change_Raw"]
        if pd.isna(val) or val is None:
            orig_change = row.get("Change %")
            if pd.notna(orig_change) and orig_change != "":
                return orig_change
            return "-"
        return f"{val:.2f}%"
        
    df_combined["Change %"] = df_combined.apply(format_change_col, axis=1)
    
    # Keep only target columns
    df_out = df_combined[["Date", "Price", "Open", "High", "Low", "Vol.", "Change %"]].copy()
    
    # Sort descending (newest first)
    df_out = df_out.iloc[::-1]
    
    # Save to CSV
    df_out.to_csv(csv_path, index=False, quoting=csv.QUOTE_NONNUMERIC)
    print(f"  Successfully updated {filename}. Total rows: {len(df_out)}")

def update_stock_indices():
    INDEX_TICKERS = {
        "FTSE Malaysia KLCI.csv":                        "^KLSE",
        "Dow Jones Industrial Average Historical Data.csv": "^DJI",
        "S&P 500.csv":                                   "^GSPC",
        "NASDAQ Composite.csv":                          "^IXIC",
        "FTSE 100 Historical Data.csv":                  "^FTSE",
        "Euro Stoxx 50 Historical Data.csv":             "^STOXX50E",
        "Nikkei 225.csv":                                "^N225",
        "Hang Seng Index.csv":                           "^HSI",
        "KOSPI Historical Data.csv":                     "^KS11",
        "Shanghai Composite.csv":                        "000001.SS",
        "TSX Composite.csv":                             "^GSPTSE",
        "TSX Venture Composite.csv":                     "^CDNX",
    }
    
    for filename, ticker in INDEX_TICKERS.items():
        try:
            update_stock_index_csv(filename, ticker)
        except Exception as e:
            print(f"Error updating {filename}: {e}")
            


def main():
    print("=== STARTING DAILY EXCHANGE RATES UPDATE ===")
    update_exchange_rates()
    print("\n=== STARTING MONTHLY MONEY SUPPLY UPDATE ===")
    update_money_supply()
    print("\n=== STARTING DAILY STOCK INDICES UPDATE ===")
    update_stock_indices()
    print("=== ALL UPDATES COMPLETED ===")

if __name__ == "__main__":
    main()

