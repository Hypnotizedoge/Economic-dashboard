"""
data_loader.py — Loads parquet data from local /data folder.
"""
import streamlit as st
import pandas as pd
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"

# ── Label Mappings ─────────────────────────────────────────────────────────────
GDP_SUPPLY_SECTORS = {
    "p0": "Overall GDP", "p1": "Agriculture", "p2": "Mining & Quarrying",
    "p3": "Manufacturing", "p4": "Construction", "p5": "Services", "p6": "Import Duties",
}
GDP_DEMAND_TYPES = {
    "e0": "Overall GDP", "e1": "Private Consumption", "e2": "Government Consumption",
    "e3": "Gross Fixed Capital Formation", "e4": "Change in Inventories",
    "e5": "Exports", "e6": "Imports",
}
CPI_DIVISIONS = {
    "overall": "Overall", "01": "Food & Beverages", "02": "Alcohol & Tobacco",
    "03": "Clothing & Footwear", "04": "Housing & Utilities", "05": "Furnishings",
    "06": "Health", "07": "Transport", "08": "Communication",
    "09": "Recreation & Culture", "10": "Education", "11": "Restaurants & Hotels",
    "12": "Miscellaneous", "13": "Food at Home",
}
PPI_SECTIONS = {
    "A": "Agriculture", "B": "Mining & Quarrying", "C": "Manufacturing",
    "D": "Electricity & Gas", "E": "Water Supply",
}
IPI_SECTIONS = {"B": "Mining", "C": "Manufacturing", "D": "Electricity"}
IOWRT_DIVISIONS = {"45": "Motor Vehicles", "46": "Wholesale", "47": "Retail"}
BOP_ACCOUNTS = {
    "ca": "Current Account", "fa": "Financial Account", "ka": "Capital Account",
    "neo": "Net Errors & Omissions", "reserves": "Reserves",
}
SITC_SECTIONS = {
    "overall": "Overall", "0": "Food", "1": "Beverages & Tobacco",
    "2": "Crude Materials", "3": "Mineral Fuels", "4": "Animal & Vegetable Oils",
    "5": "Chemicals", "6": "Manufactured Goods", "7": "Machinery & Transport",
    "8": "Misc. Manufactured", "9": "Other",
}


@st.cache_data(show_spinner=False)
def load(name: str) -> pd.DataFrame:
    """Load a local parquet file by name."""
    df = pd.read_parquet(DATA_DIR / f"{name}.parquet")
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"])
    return df


@st.cache_data(show_spinner=False)
def load_historical_gdp() -> pd.DataFrame:
    """Load the Nominal GDP 1991-2014 Excel file."""
    xlsx = Path(__file__).parent / "Nominal GDP 1991-2014.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["My Series"]
    dates, values = [], []
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                dates.append(pd.to_datetime(row[0]))
                values.append(float(row[1]))
            except (ValueError, TypeError):
                continue
    return pd.DataFrame({"date": dates, "gdp_nominal": values}).sort_values("date").reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_fx_rates() -> pd.DataFrame:
    """Load exchange rates from local FXrate.csv."""
    csv_path = Path(__file__).parent / "FXrate.csv"
    df = pd.read_csv(csv_path)
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], dayfirst=True)
    return df



@st.cache_data(show_spinner=False)
def load_money_supply() -> pd.DataFrame:
    """Load Money Supply (M1, M2, M3) from Excel."""
    xlsx = Path(__file__).parent / "Money supply.xlsx"
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["1.3"]
    dates = []
    m1_vals = []
    m2_vals = []
    m3_vals = []
    
    current_year = None
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        val_year = row[0]
        val_month = row[1]
        
        if val_year is not None:
            try:
                current_year = int(val_year)
            except (ValueError, TypeError):
                pass
                
        if val_month is not None and current_year is not None:
            try:
                month = int(val_month)
                dt = pd.to_datetime(f"{current_year}-{month:02d}-01")
                
                m3 = float(row[3]) if row[3] is not None else None
                m2 = float(row[4]) if row[4] is not None else None
                m1 = float(row[5]) if row[5] is not None else None
                
                dates.append(dt)
                m1_vals.append(m1)
                m2_vals.append(m2)
                m3_vals.append(m3)
            except (ValueError, TypeError):
                continue
                
    return pd.DataFrame({
        "date": dates,
        "m1": m1_vals,
        "m2": m2_vals,
        "m3": m3_vals
    }).sort_values("date").reset_index(drop=True)


def fseries(df: pd.DataFrame, series: str = "abs") -> pd.DataFrame:
    """Filter by series column."""
    if "series" in df.columns:
        return df[df["series"] == series].copy()
    return df.copy()


def latest(df: pd.DataFrame, col: str):
    """Return (value, date) for the most recent row."""
    if df.empty:
        return None, None
    row = df.sort_values("date").iloc[-1]
    return row[col], row["date"]


def delta(df: pd.DataFrame, col: str):
    """Return change from previous period."""
    if len(df) < 2:
        return None
    s = df.sort_values("date")
    c, p = s[col].iloc[-1], s[col].iloc[-2]
    return c - p if pd.notna(c) and pd.notna(p) else None


def fmt(value, prefix="", suffix="", d=1):
    """Smart number format."""
    if value is None:
        return "N/A"
    a = abs(value)
    if a >= 1e12:   return f"{prefix}{value/1e12:,.{d}f}T{suffix}"
    if a >= 1e9:    return f"{prefix}{value/1e9:,.{d}f}B{suffix}"
    if a >= 1e6:    return f"{prefix}{value/1e6:,.{d}f}M{suffix}"
    if a >= 1e3:    return f"{prefix}{value/1e3:,.{d}f}K{suffix}"
    return f"{prefix}{value:,.{d}f}{suffix}"
